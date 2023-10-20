import tkinter as tk
from tkinter import simpledialog, messagebox
from tkinter import ttk
import os
import openpyxl
from datetime import datetime


class PuntoDeVentaApp:
    def __init__(self):
        self.root = tk.Tk()
        self.root.title("Punto de Venta")
        self.productos = self.cargar_productos()  # Esta línea está en el constructor __init__.
        self.logged_in = False  # Variable para rastrear si se ha iniciado sesión

    def cargar_productos(self):
        try:
            archivo_productos = "productos.xlsx"
            if not os.path.exists(archivo_productos):
                return []

            workbook = openpyxl.load_workbook(archivo_productos)
            sheet = workbook.active
            productos = []

            for row in sheet.iter_rows(min_row=2, values_only=True):
                nombre, precio, cantidad = row
                producto = {"nombre": nombre, "precio": precio, "cantidad": cantidad}
                productos.append(producto)

            return productos
        except FileNotFoundError:
            return []

    def iniciar_sesion(self):
        login_frame = tk.Frame(self.root)
        login_frame.pack()
        tk.Label(login_frame, text="Nombre de usuario:").pack()
        self.username_entry = tk.Entry(login_frame)
        self.username_entry.pack()
        tk.Label(login_frame, text="Contraseña:").pack()
        self.password_entry = tk.Entry(login_frame, show="*")
        self.password_entry.pack()
        login_button = tk.Button(login_frame, text="Iniciar Sesión", command=self.check_login)
        login_button.pack()

    def check_login(self):
        username = self.username_entry.get()
        password = self.password_entry.get()

        if username == "admin" and password == "adminpass":
            dinero_inicial = simpledialog.askfloat("Dinero Inicial", "Por favor, ingresa el efectivo inicial en caja: $")
            if dinero_inicial is not None:
                self.dinero_inicial = dinero_inicial
                self.logged_in = True  # Se ha iniciado sesión
                self.admin_panel()
            else:
                messagebox.showwarning("Error", "Debes ingresar un valor válido para el dinero inicial.")
        elif username == "empleado" and password == "empleadopass":
            dinero_inicial = simpledialog.askfloat("Dinero Inicial", "Por favor, ingresa el efectivo inicial en caja: $")
            if dinero_inicial is not None:
                self.dinero_inicial = dinero_inicial
                self.logged_in = True  # Se ha iniciado sesión
                self.empleado_panel()
            else:
                messagebox.showwarning("Error", "Debes ingresar un valor válido para el dinero inicial.")
        else:
            messagebox.showerror("Error", "Credenciales incorrectas. Inténtalo de nuevo.")

    def admin_panel(self, dinero_inicial=None):
        admin_window = tk.Toplevel(self.root)
        admin_window.title("Panel de Administrador")

        def on_closing():
            admin_window.destroy()  # Cierra la ventana de administrador

        admin_window.protocol("WM_DELETE_WINDOW", on_closing)  # Configura el controlador de cierre

        tk.Button(admin_window, text="Agregar Producto", command=self.agregar_producto).pack()
        tk.Button(admin_window, text="Eliminar Producto", command=self.eliminar_producto).pack()
        tk.Button(admin_window, text="Modificar Producto", command=self.modificar_producto).pack()
        tk.Button(admin_window, text="Caja de Cobro", command=self.caja_de_cobro).pack()
        tk.Button(admin_window, text="Guardar Cambios y Salir", command=self.guardar_productos).pack()

        admin_window.mainloop()

    def agregar_producto(self):
        nombre = simpledialog.askstring("Agregar Producto", "Nombre del producto:")
        precio = simpledialog.askfloat("Agregar Producto", "Precio de venta:")
        cantidad = simpledialog.askinteger("Agregar Producto", "Cantidad en inventario:")

        if nombre and precio is not None and cantidad is not None:
            producto = {"nombre": nombre, "precio": precio, "cantidad": cantidad}
            self.productos.append(producto)
            self.guardar_productos()
            messagebox.showinfo("Éxito", f"{nombre} ha sido agregado al inventario.")
        else:
            messagebox.showwarning("Error", "Por favor, ingresa valores válidos.")

    def eliminar_producto(self):
        if not self.productos:
            messagebox.showinfo("Información", "No hay productos para eliminar.")
            return

        eliminar_window = tk.Toplevel()
        eliminar_window.title("Eliminar Producto")

        producto_names = [producto["nombre"] for producto in self.productos]
        combo = tk.StringVar()
        combo.set(producto_names[0])  # Valor inicial
        combo_box = tk.OptionMenu(eliminar_window, combo, *producto_names)
        combo_box.pack()

        def delete_product():
            selected_product = combo.get()
            if selected_product:
                for producto in self.productos:
                    if producto["nombre"] == selected_product:
                        self.productos.remove(producto)
                        self.guardar_productos()
                        messagebox.showinfo("Éxito", f"{selected_product} ha sido eliminado del inventario.")
                        eliminar_window.destroy()
                        break

        delete_button = tk.Button(eliminar_window, text="Eliminar Producto", command=delete_product)
        delete_button.pack()

    def modificar_producto(self):
        if not self.productos:
            messagebox.showinfo("Información", "No hay productos para modificar.")
            return

        modificar_window = tk.Toplevel()
        modificar_window.title("Modificar Producto")

        producto_names = [producto["nombre"] for producto in self.productos]
        combo = tk.StringVar()
        combo.set(producto_names[0])  # Valor inicial
        combo_box = tk.OptionMenu(modificar_window, combo, *producto_names)
        combo_box.pack()

        tk.Label(modificar_window, text="Nuevo Nombre:").pack()
        name_entry = tk.Entry(modificar_window)
        name_entry.pack()

        tk.Label(modificar_window, text="Nuevo Precio:").pack()
        price_entry = tk.Entry(modificar_window)
        price_entry.pack()

        tk.Label(modificar_window, text="Nuevo Inventario:").pack()
        inventory_entry = tk.Entry(modificar_window)
        inventory_entry.pack()

        def apply_changes():
            selected_product = combo.get()
            modified_name = name_entry.get()
            modified_price = price_entry.get()
            modified_inventory = inventory_entry.get()

            if selected_product and modified_name and modified_price and modified_inventory:
                for producto in self.productos:
                    if producto["nombre"] == selected_product:
                        producto["nombre"] = modified_name
                        producto["precio"] = float(modified_price)
                        producto["cantidad"] = int(modified_inventory)
                        self.guardar_productos()
                        messagebox.showinfo("Éxito", f"{selected_product} ha sido modificado en el inventario.")
                        modificar_window.destroy()
                        break

        apply_button = tk.Button(modificar_window, text="Aplicar Cambios", command=apply_changes)
        apply_button.pack()

    def guardar_productos(self):
        archivo_productos = "productos.xlsx"
        workbook = openpyxl.Workbook()
        sheet = workbook.active

        # Encabezados
        sheet.append(["Nombre", "Precio de Venta", "Cantidad en Inventario"])

        # Agrega los productos al archivo Excel
        for producto in self.productos:
            sheet.append([producto["nombre"], producto["precio"], producto["cantidad"]])

        workbook.save(archivo_productos)

    def caja_de_cobro(self):
        caja = []

        while True:
            opciones = ["Agregar producto a la caja", "Eliminar producto de la caja", "Hacer corte de caja", "Cobrar productos y volver a la caja", "Salir de la caja"]
            opcion = messagebox.askquestion("Caja de Cobro", "Opciones de caja de cobro:\n" + "\n".join(opciones))

            if opcion == "yes":
                # Opción de "Agregar producto a la caja"
                if self.productos:
                    producto_names = [producto["nombre"] for producto in self.productos]
                    selected_product = messagebox.askquestion("Agregar Producto", "Productos disponibles:\n" + "\n".join(producto_names))
                    if selected_product in producto_names:
                        index = producto_names.index(selected_product)
                        caja.append(self.productos.pop(index))
                        messagebox.showinfo("Éxito", f"{selected_product} ha sido agregado a la caja.")
                    else:
                        messagebox.showwarning("Error", "Producto no encontrado.")
                else:
                    messagebox.showinfo("Información", "No hay productos disponibles.")
            elif opcion == "no":
                # Opción de "Eliminar producto de la caja"
                if not caja:
                    messagebox.showinfo("Información", "La caja está vacía.")
                else:
                    selected_product = messagebox.askquestion("Eliminar Producto", "Productos en la caja:\n" + "\n".join([producto["nombre"] for producto in caja]))
                    if selected_product in [producto["nombre"] for producto in caja]:
                        index = [producto["nombre"] for producto in caja].index(selected_product)
                        removed_product = caja.pop(index)
                        self.productos.append(removed_product)  # Devuelve el producto al inventario
                        messagebox.showinfo("Éxito", f"{removed_product['nombre']} ha sido eliminado de la caja.")
                    else:
                        messagebox.showwarning("Error", "Producto no encontrado in the caja.")
            elif opcion == "cancel":
                # Opción de "Hacer corte de caja"
                total_ventas = self.mostrar_ventas_dia()
                dinero_en_caja = simpledialog.askfloat("Corte de Caja", "Ingrese la cantidad de dinero en caja: $")
                if dinero_en_caja is not None:
                    diferencia = dinero_en_caja - self.dinero_inicial_caja - total_ventas
                    messagebox.showinfo("Corte de Caja", f"Diferencia entre caja y ventas: ${diferencia:.2f}")
            elif opcion == "ok":
                # Opción de "Cobrar productos y volver a la caja"
                total_venta, metodo_pago = self.cobrar(caja)
                self.guardar_venta(total_venta, metodo_pago, caja)
                caja = []
            elif opcion == "Abort":
                # Opción de "Salir de la caja"
                break

    def mostrar_ventas_dia(self):
        fecha_actual = datetime.now().strftime("%Y-%m-d")
        ventas_del_dia = self.cargar_ventas(fecha_actual)

        if not ventas_del_dia:
            messagebox.showinfo("Información", "No hay ventas registradas para el día de hoy.")
            return 0

        total_ventas_dia = sum(venta[3] for venta in ventas_del_dia)

        venta_str = ""
        for venta in ventas_del_dia:
            venta_str += f"ID Venta: {venta[1]}\nProductos: {venta[2]}\nTotal Venta: ${venta[3]:.2f}\nMétodo de Pago: {venta[4]}\n\n"

        messagebox.showinfo(f"Ventas del día ({fecha_actual})", venta_str + f"Dinero total de las ventas: ${total_ventas_dia:.2f}")
        return total_ventas_dia

    def cargar_ventas(self, fecha):
        try:
            archivo_ventas = "ventas_totales.xlsx"
            if not os.path.exists(archivo_ventas):
                return []

            workbook = openpyxl.load_workbook(archivo_ventas)
            sheet = workbook.active
            ventas = []

            for row in sheet.iter_rows(min_row=2, values_only=True):
                if row[0] == fecha:
                    ventas.append(row)

            return ventas
        except FileNotFoundError:
            return []

    def guardar_venta(self, total, metodo_pago, caja):
        if not caja:
            messagebox.showwarning("Error", "No se puede guardar una venta vacía.")
            return

        fecha_actual = datetime.now().strftime("%Y-%m-d")
        archivo_ventas = "ventas_totales.xlsx"

        if not os.path.exists(archivo_ventas):
            workbook = openpyxl.Workbook()
            sheet = workbook.active
            sheet.append(["Fecha", "ID Venta", "Productos", "Total Venta", "Método de Pago"])
        else:
            workbook = openpyxl.load_workbook(archivo_ventas)
            sheet = workbook.active

        productos_vendidos = ", ".join(producto["nombre"] for producto in caja)
        nueva_fila = [fecha_actual, len(sheet["A"]) + 1, productos_vendidos, total, metodo_pago]
        sheet.append(nueva_fila)

        workbook.save(archivo_ventas)

        self.ventas.append(nueva_fila)

    def cobrar(self, caja):
        if not caja:
            messagebox.showwarning("Error", "La caja está vacía. No se puede realizar el cobro.")
            return 0, "n/a"

        total = sum(producto["precio"] for producto in caja)
        metodo_pago = simpledialog.askstring("Cobrar", f"Total a cobrar: ${total}\nMétodo de pago (efectivo/tarjeta):", initialvalue="efectivo")

        if metodo_pago:
            if metodo_pago == "efectivo":
                pago_efectivo = simpledialog.askfloat("Cobrar", "Monto en efectivo:")
                if pago_efectivo is not None:
                    cambio = pago_efectivo - total
                    if cambio < 0:
                        messagebox.showwarning("Error", "El monto en efectivo es insuficiente.")
                    else:
                        messagebox.showinfo("Cobro Realizado", f"¡Cambio: ${cambio:.2f}")
                    return total, "efectivo"
            elif metodo_pago == "tarjeta":
                messagebox.showinfo("Cobro Realizado", "Por favor, verifique que el cobro se haya efectuado de manera correcta.")
                return total, "tarjeta"
            else:
                messagebox.showwarning("Error", "Método de pago inválido.")
        return 0, "n/a"

    def empleado_panel(self):
        empleado_window = tk.Toplevel(self.root)
        empleado_window.title("Panel de Empleado")

        tk.Button(empleado_window, text="Caja de Cobro", command=self.caja_de_cobro).pack()

        empleado_window.mainloop()

if __name__ == "__main__":
    app = PuntoDeVentaApp()
    app.iniciar_sesion()
    app.root.mainloop()