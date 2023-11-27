import tkinter as tk
from tkinter import messagebox, simpledialog, ttk, PhotoImage
import customtkinter as ctk
import customtkinter
from customtkinter import CTkImage
import openpyxl
import os, json, hashlib, sys
from datetime import datetime
from PIL import Image, ImageTk
from escpos.printer import Usb
import win32print
import win32ui

#necesario para recourses
# Determinar si estamos en el entorno congelado de PyInstaller
if getattr(sys, 'frozen', False):
    # Si es así, la ruta al directorio de trabajo es el directorio del sistema donde está el ejecutable
    application_path = os.path.dirname(sys.executable)
else:
    # De lo contrario, estamos en un entorno de desarrollo normal y usamos la ruta del script
    application_path = os.path.dirname(__file__)

# Construir la ruta al directorio 'resources' relativa al entorno de trabajo
resources_path = os.path.join(application_path, 'resources')

# Utilizar resources_path para acceder a archivos dentro de la carpeta 'resources'
some_file_path = os.path.join(resources_path, 'some_file.ext')


def imprimir_ticket_windows(venta):
    nombre_impresora = "POS-58"  # Reemplaza con el nombre exacto de tu impresora
    hprinter = win32print.OpenPrinter(nombre_impresora)
    try:
        hjob = win32print.StartDocPrinter(hprinter, 1, ("Ticket de Venta", None, "RAW"))
        try:
            win32print.StartPagePrinter(hprinter)
            
            # Comando para tamaño de letra doble
            cmd_tamano_moderado = "\x1D\x21\x01"  # Ajusta según necesidades y compatibilidad de la impresora

            # Encabezado del ticket
            encabezado = "*** Ice & Crepes ***\n"
            encabezado += "Av. Hector Caballero #406,\n Valles de Huinala\n"
            encabezado += "-------------------------\n"
            
            # Datos de la venta
            info_venta = f"Fecha: {venta['fecha']}\n"
            info_venta += "Ticket No.: {:04d}\n".format(venta['ticket_numero'])
            info_venta += "-------------------------\n"
            
            # Detalle de los productos
            detalle_productos = "Producto             P.Unit\n"
            detalle_productos += "-------------------------\n"
            for producto in venta['productos']:
                detalle_productos += f"{producto['cantidad']:>4} {producto['nombre'][:15]:<15} {producto['precio']:>6.2f}\n"
            
            # Total de la venta
            total_venta = "-------------------------\n"
            total_venta += f"Total a pagar: ${venta['total']:.2f}\n"
            total_venta += "Gracias por su compra!\n"
            
            # Pie del ticket
            pie_ticket = "-------------------------\n"
            pie_ticket += "*** Vuelva Pronto! ***\n\n\n\n\n\n"
            
            # Concatenación de las partes del ticket
            ticket_completo = cmd_tamano_moderado + encabezado + info_venta + detalle_productos + total_venta + pie_ticket
            
            # Codificar el ticket a bytes y enviar a imprimir
            bytes_ticket = ticket_completo.encode('utf-8')
            win32print.WritePrinter(hprinter, bytes_ticket)
            win32print.EndPagePrinter(hprinter)
        finally:
            win32print.EndDocPrinter(hprinter)
    finally:
        win32print.ClosePrinter(hprinter)


# Variables globales
productos = []
ventas = []
tipo_usuario = None
dinero_inicial_caja = 0
entry_buscar = None  # Definir aquí para que sea accesible en todo el código
combo_productos = None  # Definir aquí para que sea accesible en todo el código
ventana_principal = None  # Declaración global
RUTA_SCRIPT = os.path.dirname(os.path.realpath(__file__))
RUTA_RESOURCES = os.path.join(RUTA_SCRIPT, "resources")

customtkinter.set_appearance_mode("light")  # Modes: system (default), light, dark
customtkinter.set_default_color_theme("green")  # Themes: blue (default), dark-blue, green
ctk.set_widget_scaling(1.25)

#funcion para centrar las ventanas
def centrar_ventana(ventana):
    ventana.update()
    ventana.update_idletasks()
    ancho_ventana = ventana.winfo_width()
    alto_ventana = ventana.winfo_height()
    x_pos = ventana.winfo_screenwidth() // 2 - ancho_ventana // 2
    y_pos = ventana.winfo_screenheight() // 2 - alto_ventana // 2
    ventana.geometry(f"+{x_pos}+{y_pos}")

def cargar_contraseñas():
    ruta_json = os.path.join(RUTA_RESOURCES, "usuarios.json")
    try:
        with open(ruta_json, "r") as file:
            return json.load(file)
    except (FileNotFoundError, json.JSONDecodeError):
        return {"admin": cifrar_contrasena("12345"), "empleado": cifrar_contrasena("12345")}

def guardar_contraseñas(usuarios):
    ruta_json = os.path.join(RUTA_RESOURCES, "usuarios.json")
    with open(ruta_json, "w") as file:
        json.dump(usuarios, file)


def actualizar_contraseñas(contrasena_admin, contrasena_empleado):
    usuarios = cargar_contraseñas()
    usuarios["admin"] = cifrar_contrasena(contrasena_admin)
    usuarios["empleado"] = cifrar_contrasena(contrasena_empleado)
    guardar_contraseñas(usuarios)
    messagebox.showinfo("Información", "Contraseñas actualizadas correctamente")

def cifrar_contrasena(contrasena):
    return hashlib.sha256(contrasena.encode()).hexdigest()


def manejar_dinero_inicial():
    ruta_archivo = os.path.join(RUTA_RESOURCES, "dinero_inicial.json")
    fecha_actual = datetime.now().strftime("%Y-%m-%d")

    # Verifica si el archivo ya existe
    if os.path.isfile(ruta_archivo):
        with open(ruta_archivo, 'r') as archivo:
            try:
                datos = json.load(archivo)
                # Verifica si la fecha guardada es la actual
                if datos.get('fecha') == fecha_actual:
                    return datos.get('dinero_inicial')
            except json.JSONDecodeError:
                # Manejo de un posible error de decodificación JSON (archivo vacío o malformado)
                pass

    # Pide el dinero inicial y lo guarda si no hay datos válidos para la fecha actual
    dinero_inicial = float(simpledialog.askstring("Dinero Inicial", "Por favor, ingresa el efectivo inicial en caja: $"))
    with open(ruta_archivo, 'w') as archivo:
        json.dump({'fecha': fecha_actual, 'dinero_inicial': dinero_inicial}, archivo)
    return dinero_inicial

# Funciones para interactuar con Excel
def guardar_productos():
    ruta_archivo = os.path.join(RUTA_RESOURCES, "productos.xlsx")
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.append(["Nombre", "Precio de Venta", "Cantidad en Inventario"])
    for producto in productos:
        sheet.append([producto["nombre"], producto["precio"], producto["cantidad"]])
    workbook.save(ruta_archivo)
    
def cargar_productos():
    ruta_archivo = os.path.join(RUTA_RESOURCES, "productos.xlsx")
    try:
        workbook = openpyxl.load_workbook(ruta_archivo)
        sheet = workbook.active
        for row in sheet.iter_rows(min_row=2, values_only=True):
            productos.append({"nombre": row[0], "precio": row[1], "cantidad": row[2]})
    except FileNotFoundError:
        print("Archivo de productos no encontrado.")

def guardar_venta(venta):
    ruta_archivo_json = os.path.join(RUTA_SCRIPT, "ventas.json")
    ruta_archivo_xlsx = os.path.join(RUTA_SCRIPT, "ventas_totales.xlsx")
    ventas.append(venta)

    with open(ruta_archivo_json, 'w') as file:
        json.dump(ventas, file)

    if not os.path.exists(ruta_archivo_xlsx):
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.append(["Fecha", "ID Venta", "Productos", "Total Venta", "Método de Pago"])
    else:
        workbook = openpyxl.load_workbook(ruta_archivo_xlsx)
        sheet = workbook.active

    sheet.append(venta)
    workbook.save(ruta_archivo_xlsx)

def cargar_ventas():
    global ventas
    ruta_archivo = os.path.join(RUTA_SCRIPT, "ventas.json")
    try:
        with open(ruta_archivo, 'r') as file:
            ventas = json.load(file)
    except (FileNotFoundError, json.JSONDecodeError):
        ventas = []


# Inicio de Sesión
def iniciar_sesion(usuario, contrasena, ventana_login):
    usuarios = cargar_contraseñas()
    contrasena_cifrada = cifrar_contrasena(contrasena)
    if usuario in usuarios and usuarios[usuario] == contrasena_cifrada:
        global tipo_usuario
        tipo_usuario = usuario
        ventana_login.destroy()
        
        # Llama a manejar_dinero_inicial en lugar de simpledialog.askstring
        global dinero_inicial_caja
        dinero_inicial_caja = manejar_dinero_inicial()
        
        if tipo_usuario == "admin":
            mostrar_ventana_principal()
        elif tipo_usuario == "empleado":
            mostrar_ventana_empleado()
    else:
        messagebox.showerror("Error", "Credenciales incorrectas")

   

def ventana_modificar_contrasenas():
    ventana_modificar = ctk.CTkToplevel()
    ventana_modificar.title("Modificar Contraseñas")
    centrar_ventana(ventana_modificar)

    ventana_modificar.attributes('-topmost', 1)  # Configura la ventana como topmost
    ventana_modificar.grab_set()  # Hace que la ventana sea modal
    

    frame = ctk.CTkFrame(ventana_modificar)
    frame.grid(row=0, column=0, sticky=("nsew"))

    ctk.CTkLabel(frame, text="Nueva Contraseña Admin:").grid(row=0, column=0, sticky="w")
    entry_admin = ctk.CTkEntry(frame)
    entry_admin.grid(row=0, column=1, sticky="ew")

    ctk.CTkLabel(frame, text="Nueva Contraseña Empleado:").grid(row=1, column=0, sticky="w")
    entry_empleado = ctk.CTkEntry(frame)
    entry_empleado.grid(row=1, column=1, sticky="ew")

    ctk.CTkButton(frame, text="Actualizar", command=lambda: actualizar_contraseñas(entry_admin.get(), entry_empleado.get())).grid(row=2, column=1, sticky="e")

    for child in frame.winfo_children():
        child.grid_configure(padx=5, pady=5)

def on_entry_click(event, default_text, entry):
    """Función que se llama cuando se hace clic en el Entry."""
    if entry.get() == default_text:
        entry.delete(0, "end")  # Borra el placeholder
        entry.configure(fg_color="white")  # Cambia el color del texto

def on_focusout(event, default_text, entry):
    """Función que se llama cuando el Entry pierde el foco."""
    if entry.get().strip() == '':
        entry.insert(0, default_text)
        entry.configure(fg_color="#white")  # Cambia el color del texto a gris

def mostrar_ventana_login():
    global entry_usuario, entry_contrasena
    ventana_login = ctk.CTk()
    ventana_login.title("Inicio de Sesión")
    ventana_login.geometry("800x400")
    centrar_ventana(ventana_login)

    ventana_login.attributes('-topmost', 1)  # Configura la ventana como topmost
    ventana_login.grab_set()  # Hace que la ventana sea modal

    # Configura el grid de la ventana para que los elementos se expandan proporcionalmente
    ventana_login.grid_columnconfigure(0, weight=2)  # Aumentar para empujar los widgets a la derecha
    ventana_login.grid_columnconfigure(1, weight=0)  # Columna de la línea no se expande
    ventana_login.grid_columnconfigure(2, weight=1)  # Reducir para mover los widgets a la derecha
    ventana_login.grid_rowconfigure(0, weight=1)

    # Frame para la imagen del lado izquierdo
    frame_imagen = ctk.CTkFrame(ventana_login)
    frame_imagen.grid(row=0, column=0, sticky='nsew')
    frame_imagen.grid_rowconfigure(0, weight=1)
    frame_imagen.grid_columnconfigure(0, weight=1)

    # Cargar y mostrar la imagen
    ruta_imagen = os.path.join(RUTA_RESOURCES, "10222.png")
    imagen = Image.open(ruta_imagen)
    imagen = imagen.resize((350, 400), Image.Resampling.LANCZOS)
    foto = ImageTk.PhotoImage(imagen)
    label_imagen = ctk.CTkLabel(frame_imagen, image=foto, text="")
    label_imagen.grid(row=0, column=0, sticky='nsew')

    # Canvas para la línea divisoria
    canvas = ctk.CTkCanvas(ventana_login, width=2, height=400)
    canvas.grid(row=0, column=1, sticky="ns")
    canvas.create_line(1, 0, 1, 400, fill="gray")

    # Frame central donde estarán los widgets de login
    frame_central = ctk.CTkFrame(ventana_login)
    frame_central.grid(row=0, column=2, sticky='nsew')
    frame_central.grid_rowconfigure(1, weight=1)
    frame_central.grid_rowconfigure(5, weight=2)
    frame_central.grid_columnconfigure(0, weight=1)
    frame_central.grid_columnconfigure(1, weight=1)

    # Título de la sección de inicio de sesión
    titulo_sesion = ctk.CTkLabel(frame_central, text="Iniciar Sesión", font=("Roboto", 24))
    titulo_sesion.grid(row=0, column=0, columnspan=2, pady=(20, 20))

    # Cargar y mostrar el icono TENGO QUE CAMBIARLO
    ruta_icono = os.path.join(RUTA_RESOURCES, "usuario.png")  # Cambia esto a la ruta de tu icono
    imagen_icono = Image.open(ruta_icono)
    imagen_icono = imagen_icono.resize((50, 50), Image.Resampling.LANCZOS)  # Ajusta el tamaño según necesites
    foto_icono = ImageTk.PhotoImage(imagen_icono)
    label_icono = ctk.CTkLabel(frame_central, image=foto_icono, text="")
    label_icono.grid(row=1, column=0, columnspan=2)  # Ajusta la fila según necesites
    label_icono.image = foto_icono  # Guarda una referencia

    # Cargar y mostrar el icono de usuario
    ruta_icono_usuario = os.path.join(RUTA_RESOURCES, "grupo.png")  # Cambia esto a la ruta de tu icono de usuario
    imagen_icono_usuario = Image.open(ruta_icono_usuario)
    imagen_icono_usuario = imagen_icono_usuario.resize((30, 30), Image.Resampling.LANCZOS)  # Ajusta el tamaño según necesites
    foto_icono_usuario = ImageTk.PhotoImage(imagen_icono_usuario)
    label_icono_usuario = ctk.CTkLabel(frame_central, image=foto_icono_usuario, text="")
    label_icono_usuario.grid(row=2, column=0, padx = (10, 0))
    label_icono_usuario.image = foto_icono_usuario  # Guarda una referencia
    
    # Campo de entrada para el usuario
    placeholder_usuario = "Usuario"
    entry_usuario = ctk.CTkEntry(frame_central)
    entry_usuario.insert(0, placeholder_usuario)
    entry_usuario.configure(cursor="xterm")
    entry_usuario.configure(fg_color="white")
    entry_usuario.bind("<FocusIn>", lambda event: on_entry_click(event, placeholder_usuario, entry_usuario))
    entry_usuario.bind("<FocusOut>", lambda event: on_focusout(event, placeholder_usuario, entry_usuario))
    entry_usuario.grid(row=2, column=1, sticky="w", padx=(0,20))

    # Cargar y mostrar el icono de contraseña
    ruta_icono_contrasena = os.path.join(RUTA_RESOURCES, "candado.png")  # Cambia esto a la ruta de tu icono de contraseña
    imagen_icono_contrasena = Image.open(ruta_icono_contrasena)
    imagen_icono_contrasena = imagen_icono_contrasena.resize((30, 30), Image.Resampling.LANCZOS)  # Ajusta el tamaño según necesites
    foto_icono_contrasena = ImageTk.PhotoImage(imagen_icono_contrasena)
    label_icono_contrasena = ctk.CTkLabel(frame_central, image=foto_icono_contrasena, text="")
    label_icono_contrasena.grid(row=3, column=0, padx =(10, 0))
    label_icono_contrasena.image = foto_icono_contrasena  # Guarda una referencia

    # Campo de entrada para la contraseña
    placeholder_contrasena = "Contraseña"
    entry_contrasena = ctk.CTkEntry(frame_central, show="*")
    entry_contrasena.insert(0, placeholder_contrasena)
    entry_contrasena.configure(cursor="xterm")
    entry_contrasena.configure(fg_color="white")
    entry_contrasena.bind("<FocusIn>", lambda event: on_entry_click(event, placeholder_contrasena, entry_contrasena))
    entry_contrasena.bind("<FocusOut>", lambda event: on_focusout(event, placeholder_contrasena, entry_contrasena))
    entry_contrasena.grid(row=3, column=1, sticky="w", padx=(0, 20))

    boton_login = ctk.CTkButton(frame_central, text="Iniciar sesión", command=lambda: iniciar_sesion(entry_usuario.get(), entry_contrasena.get(), ventana_login))
    boton_login.grid(row=4, column=0, columnspan=2, padx=(15,0), pady=10)

    # Para mantener la referencia de la imagen
    label_imagen.image = foto

    ventana_login.mainloop()


# Ventana Principal
def mostrar_ventana_principal():
    
    ventana_principal = ctk.CTk()
    ventana_principal.title("ADMINISTRADOR")
    ventana_principal.geometry("800x400")  # Ajusta esto si necesitas una ventana de diferente tamaño
    centrar_ventana(ventana_principal)

    ventana_principal.attributes('-topmost', 1)  # Configura la ventana como topmost
    ventana_principal.grab_set()  # Hace que la ventana sea modal

    # Configura el grid de la ventana para que los elementos se expandan proporcionalmente
    ventana_principal.grid_columnconfigure(0, weight=1)
    ventana_principal.grid_rowconfigure(0, weight=1)

    frame = ctk.CTkFrame(ventana_principal)
    frame.grid(row=0, column=0, sticky="nsew")
    frame.grid_columnconfigure((0, 1), weight=1)
    frame.grid_rowconfigure((0, 1, 2), weight=1)

   # Definición de los iconos para los botones
    iconos = {
        "Agregar Producto": os.path.join(RUTA_RESOURCES, "boton-agregar.png"),
        "Eliminar Producto": os.path.join(RUTA_RESOURCES, "contenedor-de-basura.png"),
        "Modificar Producto": os.path.join(RUTA_RESOURCES, "editar.png"),
        "Modificar Contraseñas": os.path.join(RUTA_RESOURCES, "rueda-dentada.png"),
        "Caja de Cobro": os.path.join(RUTA_RESOURCES, "monitor.png"),
        "Ventas": os.path.join(RUTA_RESOURCES, "informe-de-venta.png"),
        "Salir": os.path.join(RUTA_RESOURCES, "cerrar-sesion.png")
    }
    # Define los botones y sus respectivas funciones
    botones_info = [
        ("Agregar Producto", agregar_producto),
        ("Eliminar Producto", eliminar_producto),
        ("Modificar Producto", modificar_producto),
        ("Modificar Contraseñas", ventana_modificar_contrasenas),
        ("Caja de Cobro", lambda: caja_de_cobro(ventana_principal)),
        ("Ventas", ver_ventas),
        ("Salir", ventana_principal.destroy)
    ]

    def cargar_icono_redimensionado(ruta, nuevo_ancho, nuevo_alto):
        imagen = Image.open(ruta)
        imagen_redimensionada = imagen.resize((nuevo_ancho, nuevo_alto), Image.LANCZOS)
        return ImageTk.PhotoImage(imagen_redimensionada)

    # Carga los iconos para los botones y redimensiona
    tamanio_icono = (50, 50)  # Ejemplo: 50x50 píxeles
    for i, (texto, comando) in enumerate(botones_info):
        ruta_icono = iconos.get(texto, "")
        icono = cargar_icono_redimensionado(ruta_icono, *tamanio_icono)
        boton = ctk.CTkButton(frame, text=texto, command=comando, image=icono, compound="left")
        boton.image = icono  # Guarda una referencia al icono para evitar la recolección de basura
        
        # Posicionamiento y configuración del botón
        fila = i // 2  # Dos botones por fila
        columna = i % 2  # Dos columnas
        boton.grid(row=fila, column=columna, padx=10, pady=10, sticky="nsew")

    ventana_principal.mainloop()

def mostrar_ventana_empleado():
    ventana_empleado = ctk.CTk()
    ventana_empleado.title("EMPLEADO")
    ventana_empleado.geometry("800x400")  # Ajusta esto si necesitas una ventana de diferente tamaño
    centrar_ventana(ventana_empleado)

    ventana_empleado.attributes('-topmost', 1)  # Configura la ventana como topmost
    ventana_empleado.grab_set()  # Hace que la ventana sea modal

    # Configura el grid de la ventana para que los elementos se expandan proporcionalmente
    ventana_empleado.grid_columnconfigure(0, weight=1)
    ventana_empleado.grid_rowconfigure(0, weight=1)

    frame = ctk.CTkFrame(ventana_empleado)
    frame.grid(row=0, column=0, sticky="nsew")
    frame.grid_columnconfigure(0, weight=1)
    frame.grid_rowconfigure((0, 1), weight=1)

    # Definición de los iconos para los botones
    iconos = {
        "Caja de Cobro": os.path.join(RUTA_RESOURCES, "monitor.png"),
        "Salir": os.path.join(RUTA_RESOURCES, "cerrar-sesion.png")
    }

    # Define los botones y sus respectivas funciones
    botones_info = [
        ("Caja de Cobro", lambda: caja_de_cobro(ventana_empleado)),
        ("Salir", ventana_empleado.destroy)
    ]

    def cargar_icono_redimensionado(ruta, nuevo_ancho, nuevo_alto):
        imagen = Image.open(ruta)
        imagen_redimensionada = imagen.resize((nuevo_ancho, nuevo_alto), Image.LANCZOS)
        return ImageTk.PhotoImage(imagen_redimensionada)

    # Carga los iconos para los botones y redimensiona
    tamanio_icono = (50, 50)  # Ejemplo: 50x50 píxeles
    for i, (texto, comando) in enumerate(botones_info):
        ruta_icono = iconos.get(texto, "")
        icono = cargar_icono_redimensionado(ruta_icono, *tamanio_icono)
        boton = ctk.CTkButton(frame, text=texto, command=comando, image=icono, compound="left")
        boton.image = icono  # Guarda una referencia al icono para evitar la recolección de basura
        
        # Posicionamiento y configuración del botón
        boton.grid(row=i, column=0, padx=10, pady=10, sticky="nsew")

    ventana_empleado.mainloop()


# Función para agregar producto
def agregar_producto():
    ventana_agregar = ctk.CTkToplevel()
    ventana_agregar.title("Agregar Producto")
    centrar_ventana(ventana_agregar)
    
    ventana_agregar.attributes('-topmost', 1)  # Configura la ventana como topmost
    ventana_agregar.grab_set()  # Hace que la ventana sea modal

    frame = ctk.CTkFrame(ventana_agregar)
    frame.grid(row=0, column=0, sticky=(ctk.W, ctk.E, ctk.N, ctk.S))

    ctk.CTkLabel(frame, text="Nombre del producto:").grid(column=0, row=0, sticky=ctk.W)
    entry_nombre = ctk.CTkEntry(frame)
    entry_nombre.grid(column=1, row=0, sticky=(ctk.W, ctk.E))

    ctk.CTkLabel(frame, text="Precio de venta:").grid(column=0, row=1, sticky=ctk.W)
    entry_precio = ctk.CTkEntry(frame)
    entry_precio.grid(column=1, row=1, sticky=(ctk.W, ctk.E))

    ctk.CTkLabel(frame, text="Cantidad en inventario:").grid(column=0, row=2, sticky=ctk.W)
    entry_cantidad = ctk.CTkEntry(frame)
    entry_cantidad.grid(column=1, row=2, sticky=(ctk.W, ctk.E))

    ctk.CTkButton(frame, text="Agregar", command=lambda: guardar_nuevo_producto()).grid(column=1, row=3, sticky=ctk.E)

    # Función interna para guardar el nuevo producto
    def guardar_nuevo_producto():
        nombre = entry_nombre.get()
        precio = float(entry_precio.get())
        cantidad = int(entry_cantidad.get())
        productos.append({"nombre": nombre, "precio": precio, "cantidad": cantidad})
        ventana_agregar.destroy()
        guardar_productos()

    # Configurar el espaciado y expansión de la cuadrícula
    for child in frame.winfo_children():
        child.grid_configure(padx=5, pady=5)


# Función para eliminar producto
def actualizar_option_menu(event=None):
    global combo_productos, entry_buscar, frame
    busqueda = entry_buscar.get().lower() if entry_buscar and entry_buscar.get() else ""
    productos_filtrados = [producto["nombre"] for producto in productos if busqueda in producto["nombre"].lower()]

    # Destruir el CTkOptionMenu anterior
    if combo_productos:
        combo_productos.destroy()

    # Crear un nuevo CTkOptionMenu con los productos filtrados
    combo_productos = ctk.CTkOptionMenu(frame, variable=tk.StringVar(), values=productos_filtrados)
    combo_productos.grid(column=1, row=1, sticky=(ctk.W, ctk.E))

    # Establecer el primer producto filtrado como seleccionado, si hay alguno
    if productos_filtrados:
        combo_productos.set(productos_filtrados[0])
    else:
        combo_productos.set('')

def eliminar():
    producto_seleccionado = combo_productos.get()
    indice_a_eliminar = None
    for indice, producto in enumerate(productos):
        if producto["nombre"] == producto_seleccionado:
            indice_a_eliminar = indice
            break

    if indice_a_eliminar is not None:
        del productos[indice_a_eliminar]
        guardar_productos()
        actualizar_option_menu(None)
    else:
        messagebox.showwarning("Advertencia", "Por favor, selecciona un producto")

def eliminar_producto():
    global entry_buscar, combo_productos, frame
    ventana_eliminar = ctk.CTkToplevel()
    ventana_eliminar.title("Eliminar Producto")
    centrar_ventana(ventana_eliminar)
    
    ventana_eliminar.attributes('-topmost', 1)  # Configura la ventana como topmost
    ventana_eliminar.grab_set()  # Hace que la ventana sea modal

    frame = ctk.CTkFrame(ventana_eliminar)
    frame.grid(row=0, column=0, sticky=(ctk.W, ctk.E, ctk.N, ctk.S))

    ctk.CTkLabel(frame, text="Buscar Producto:").grid(column=0, row=0, sticky=ctk.W)
    entry_buscar = ctk.CTkEntry(frame)
    entry_buscar.grid(column=1, row=0, sticky=(ctk.W, ctk.E))
    entry_buscar.bind('<KeyRelease>', actualizar_option_menu)

    # Crear el CTkOptionMenu inicial con todos los productos
    combo_productos = ctk.CTkOptionMenu(frame, variable=tk.StringVar(), values=[producto["nombre"] for producto in productos])
    combo_productos.grid(column=1, row=1, sticky=(ctk.W, ctk.E))
    if productos:
        combo_productos.set(productos[0]["nombre"])

    ctk.CTkButton(frame, text="Eliminar", command=eliminar).grid(column=1, row=2, sticky=ctk.E)

    for child in frame.winfo_children():
        child.grid_configure(padx=5, pady=5)


# Función para actualizar el OptionMenu
def actualizar_option_menu_modificar(event=None):
    global combo_productos_modificar, entry_buscar_modificar, frame_modificar
    busqueda = entry_buscar_modificar.get().lower() if entry_buscar_modificar and entry_buscar_modificar.get() else ""
    productos_filtrados = [producto["nombre"] for producto in productos if busqueda in producto["nombre"].lower()]

    if combo_productos_modificar:
        combo_productos_modificar.destroy()

    combo_productos_modificar = ctk.CTkOptionMenu(frame_modificar, variable=tk.StringVar(), values=productos_filtrados)
    combo_productos_modificar.grid(column=1, row=1, sticky=(ctk.W, ctk.E))

    if productos_filtrados:
        combo_productos_modificar.set(productos_filtrados[0])
    else:
        combo_productos_modificar.set('')

# Función para cargar los detalles del producto para modificar
def cargar_producto_para_modificar():
    producto_seleccionado = combo_productos_modificar.get()
    for producto in productos:
        if producto["nombre"] == producto_seleccionado:
            entry_nombre.delete(0, tk.END)
            entry_nombre.insert(0, producto["nombre"])
            entry_precio.delete(0, tk.END)
            entry_precio.insert(0, producto["precio"])
            entry_cantidad.delete(0, tk.END)
            entry_cantidad.insert(0, producto["cantidad"])
            break

# Función para guardar los cambios del producto
def guardar_cambios_producto():
    producto_seleccionado = combo_productos_modificar.get()
    for producto in productos:
        if producto["nombre"] == producto_seleccionado:
            producto["nombre"] = entry_nombre.get()
            producto["precio"] = float(entry_precio.get())
            producto["cantidad"] = int(entry_cantidad.get())
            break
    actualizar_option_menu_modificar()  # Actualizar el OptionMenu
    messagebox.showinfo("Éxito", "Producto modificado con éxito")

# Función para modificar producto
def modificar_producto():
    global entry_buscar_modificar, combo_productos_modificar, frame_modificar, entry_nombre, entry_precio, entry_cantidad
    ventana_modificar = ctk.CTkToplevel()
    ventana_modificar.title("Modificar Producto")
    centrar_ventana(ventana_modificar)
    
    ventana_modificar.attributes('-topmost', 1)  # Configura la ventana como topmost
    ventana_modificar.grab_set()  # Hace que la ventana sea modal

    frame_modificar = ctk.CTkFrame(ventana_modificar)
    frame_modificar.grid(row=0, column=0, sticky=(ctk.W, ctk.E, ctk.N, ctk.S))

    ctk.CTkLabel(frame_modificar, text="Buscar Producto:").grid(column=0, row=0, sticky=ctk.W)
    entry_buscar_modificar = ctk.CTkEntry(frame_modificar)
    entry_buscar_modificar.grid(column=1, row=0, sticky=(ctk.W, ctk.E))
    entry_buscar_modificar.bind('<KeyRelease>', actualizar_option_menu_modificar)

    combo_productos_modificar = ctk.CTkOptionMenu(frame_modificar, variable=tk.StringVar(), values=[producto["nombre"] for producto in productos])
    combo_productos_modificar.grid(column=1, row=1, sticky=(ctk.W, ctk.E))
    if productos:
        combo_productos_modificar.set(productos[0]["nombre"])

    ctk.CTkLabel(frame_modificar, text="Nombre del producto:").grid(column=0, row=2, sticky=ctk.W)
    entry_nombre = ctk.CTkEntry(frame_modificar)
    entry_nombre.grid(column=1, row=2, sticky=(ctk.W, ctk.E))

    ctk.CTkLabel(frame_modificar, text="Precio de venta:").grid(column=0, row=3, sticky=ctk.W)
    entry_precio = ctk.CTkEntry(frame_modificar)
    entry_precio.grid(column=1, row=3, sticky=(ctk.W, ctk.E))

    ctk.CTkLabel(frame_modificar, text="Cantidad en inventario:").grid(column=0, row=4, sticky=ctk.W)
    entry_cantidad = ctk.CTkEntry(frame_modificar)
    entry_cantidad.grid(column=1, row=4, sticky=(ctk.W, ctk.E))

    ctk.CTkButton(frame_modificar, text="Cargar Producto", command=cargar_producto_para_modificar).grid(column=0, row=5, sticky=ctk.W)
    ctk.CTkButton(frame_modificar, text="Guardar Cambios", command=guardar_cambios_producto).grid(column=1, row=5, sticky=ctk.E)

    for child in frame_modificar.winfo_children():
        child.grid_configure(padx=5, pady=5)

    actualizar_option_menu_modificar()

def ver_ventas():
    ventana_ver_ventas = ctk.CTkToplevel()
    ventana_ver_ventas.title("Ventas")
    centrar_ventana(ventana_ver_ventas)
    ventana_ver_ventas.geometry("800x400")
    centrar_ventana(ventana_ver_ventas)

    lista_ventas = tk.Listbox(ventana_ver_ventas)
    lista_ventas.pack(expand=True, fill='both')

    for venta in ventas:
        lista_ventas.insert(tk.END, f"ID Venta: {venta['ID']}, Fecha: {venta['fecha']}, Total: ${venta['total']}")

    boton_cancelar_venta = ctk.CTkButton(ventana_ver_ventas, text="Cancelar Venta", command=lambda: cancelar_venta(lista_ventas.curselection(), ventana_ver_ventas))
    boton_cancelar_venta.pack()

def cancelar_venta(seleccion, ventana_ver_ventas):
    if seleccion:
        indice_venta = seleccion[0]
        venta_seleccionada = ventas[indice_venta]

        # Eliminar de la lista en memoria
        del ventas[indice_venta]

        # Actualizar el archivo JSON
        with open(os.path.join(RUTA_SCRIPT, "ventas.json"), 'w') as file:
            json.dump(ventas, file)

        # Actualizar el archivo Excel
        actualizar_archivo_excel()

        ventana_ver_ventas.destroy()
        messagebox.showinfo("Cancelación", f"La venta ha sido cancelada.")
    else:
        messagebox.showwarning("Advertencia", "Por favor, selecciona una venta para cancelar")

def actualizar_archivo_excel():
    ruta_archivo_xlsx = os.path.join(RUTA_SCRIPT, "ventas_totales.xlsx")
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.append(["Fecha", "ID Venta", "Productos", "Total Venta", "Método de Pago"])

    for venta in ventas:
        sheet.append([venta['fecha'], venta['ID'], ", ".join(p['nombre'] for p in venta['productos']), venta['total'], venta['metodo_pago']])

    workbook.save(ruta_archivo_xlsx)



def caja_de_cobro(ventana_padre):
    caja = []
    ventana_cobro = ctk.CTkToplevel(ventana_padre)
    ventana_cobro.title("Caja de Cobro")
    ventana_cobro.state('zoomed')  # Pantalla completa

    frame = ctk.CTkFrame(ventana_cobro)
    frame.grid(row=0, column=0, sticky="nsew")
    ventana_cobro.grid_rowconfigure(0, weight=1)
    ventana_cobro.grid_columnconfigure(0, weight=1)

    iconos_caja_cobro = {
    "Agregar": os.path.join(RUTA_RESOURCES, "boton-agregar.png"),
    "Eliminar": os.path.join(RUTA_RESOURCES, "basura.png"),
    "Efectivo": os.path.join(RUTA_RESOURCES, "dinero.png"),
    "Tarjeta": os.path.join(RUTA_RESOURCES, "tarjeta-de-credito.png"),
    "CorteCaja": os.path.join(RUTA_RESOURCES, "salario.png")
}

    def cargar_icono_redimensionado(ruta, nuevo_ancho, nuevo_alto):
        imagen = Image.open(ruta)
        imagen_redimensionada = imagen.resize((nuevo_ancho, nuevo_alto), Image.LANCZOS)
        return ImageTk.PhotoImage(imagen_redimensionada)

    # Carga los iconos para los botones y redimensiona
    tamanio_icono_caja = (30, 30)  # Ajusta el tamaño según necesites
    icono_agregar = cargar_icono_redimensionado(iconos_caja_cobro["Agregar"], *tamanio_icono_caja)
    icono_eliminar = cargar_icono_redimensionado(iconos_caja_cobro["Eliminar"], *tamanio_icono_caja)
    icono_efectivo = cargar_icono_redimensionado(iconos_caja_cobro["Efectivo"], *tamanio_icono_caja)
    icono_tarjeta = cargar_icono_redimensionado(iconos_caja_cobro["Tarjeta"], *tamanio_icono_caja)
    icono_corte_caja = cargar_icono_redimensionado(iconos_caja_cobro["CorteCaja"], *tamanio_icono_caja)


    # Buscador de Productos
    ctk.CTkLabel(frame, text="Buscar Producto:").grid(column=0, row=0, sticky=ctk.W)
    entry_buscar = ttk.Entry(frame)
    entry_buscar.grid(column=1, row=0, sticky="ew")
    frame.grid_columnconfigure(1, weight=1)

    # Implementación de la función actualizar_lista_productos
    def actualizar_lista_productos(event):
        busqueda = entry_buscar.get().lower()
        productos_filtrados = [producto for producto in productos if busqueda in producto["nombre"].lower()]
        lista_productos.delete(0, ctk.END)
        for producto in productos_filtrados:
            lista_productos.insert(ctk.END, producto["nombre"])

    entry_buscar.bind('<KeyRelease>', actualizar_lista_productos)

    lista_productos = tk.Listbox(frame)
    lista_productos.grid(column=0, row=1, columnspan=2, sticky="nsew")
    frame.grid_rowconfigure(1, weight=1)

    # Etiqueta para mostrar el costo total
    etiqueta_costo_total = ctk.CTkLabel(frame, text="TOTAL: $0.00")
    etiqueta_costo_total.grid(column=0, row=2, sticky=ctk.W)

    # Lista de Productos en Caja con Enumeración
    lista_caja = tk.Listbox(frame)
    lista_caja.grid(column=0, row=3, columnspan=2, sticky="nsew")
    frame.grid_rowconfigure(3, weight=1)

    # Función para agregar productos a la caja
    def agregar_a_caja():
        seleccion = lista_productos.curselection()
        if seleccion:
            nombre_producto_seleccionado = lista_productos.get(seleccion[0])
            producto_seleccionado = next((producto for producto in productos if producto["nombre"] == nombre_producto_seleccionado), None)
            
            if producto_seleccionado:
                # Obtener la cantidad del producto deseada por el usuario
                cantidad = simpledialog.askinteger("Cantidad", "Ingresa la cantidad del producto:", minvalue=1, maxvalue=100)
                
                if cantidad is not None and cantidad > 0:
                    # Crear una copia del producto seleccionado con la cantidad especificada
                    producto_con_cantidad = producto_seleccionado.copy()
                    producto_con_cantidad['cantidad'] = cantidad

                    caja.append(producto_con_cantidad)
                    actualizar_lista_caja()
                    actualizar_costo_total()
                else:
                    messagebox.showwarning("Advertencia", "Cantidad no válida")

    boton_agregar = ctk.CTkButton(frame, text="Agregar a Caja", command=agregar_a_caja, image=icono_agregar, compound="left")
    boton_agregar.image = icono_agregar
    boton_agregar.grid(column=0, row=4, sticky=ctk.E)

    #funcion para eliminar productos de la caja
    def eliminar_de_caja():
        seleccion = lista_caja.curselection()
        if seleccion:
            del caja[seleccion[0]]
            actualizar_lista_caja()
            actualizar_costo_total()

    boton_eliminar = ctk.CTkButton(frame, text="Eliminar de Caja", command=eliminar_de_caja, image=icono_eliminar, compound="left")
    boton_eliminar.image = icono_eliminar
    boton_eliminar.grid(column=1, row=4, sticky=ctk.W)

    # Función para actualizar el costo total
    def actualizar_costo_total():
        total = sum(p["precio"] * p["cantidad"] for p in caja)
        etiqueta_costo_total.configure(text=f"Costo Total: ${total:.2f}")

    # Función para actualizar la lista de productos en caja
    def actualizar_lista_caja():
        lista_caja.delete(0, tk.END)
        for producto in caja:
            lista_caja.insert(tk.END, f"{producto['nombre']} x {producto['cantidad']}")

    # Funciones de Método de Pago
    def realizar_cobro(metodo_pago):
        total = sum(p["precio"] * p["cantidad"] for p in caja)
        if metodo_pago == "efectivo":
            pago_cliente = simpledialog.askfloat("Pago en Efectivo", "Cliente paga con: $", minvalue=total)
            if pago_cliente:
                cambio = pago_cliente - total
                messagebox.showinfo("Cambio", f"Cambio: ${cambio:.2f}")
        elif metodo_pago == "tarjeta":
            messagebox.showinfo("Pago", "Verificar el pago")


        fecha_actual = datetime.now().strftime("%Y-%m-%d")
         # El ID de la venta es igual al número de ventas registradas más uno
        id_venta = len(ventas) + 1
        venta = [fecha_actual, id_venta, ", ".join(p["nombre"] for p in caja), total, metodo_pago]
        guardar_venta(venta)

        venta_ticket = {
            'fecha': fecha_actual,
            'ticket_numero': id_venta,
            'productos': caja,
            'total': total
        }

        imprimir_ticket_windows(venta_ticket)  # Llamar a la función de impresión

        caja.clear()
        actualizar_lista_caja()

    # Funciones de Método de Pago y Corte de Caja
    boton_efectivo = ctk.CTkButton(frame, text="Pagar con Efectivo", command=lambda: realizar_cobro("efectivo"), image=icono_efectivo, compound="left")
    boton_efectivo.image = icono_efectivo
    boton_efectivo.grid(column=0, row=5, sticky=ctk.E, padx=(0,10))

    boton_tarjeta = ctk.CTkButton(frame, text="Pagar con Tarjeta", command=lambda: realizar_cobro("tarjeta"), image=icono_tarjeta, compound="left")
    boton_tarjeta.image = icono_tarjeta
    boton_tarjeta.grid(column=1, row=5, sticky=ctk.W)


    # Función para el corte de caja
    def corte_de_caja():
        try:
            workbook = openpyxl.load_workbook("ventas_totales.xlsx")
            sheet = workbook.active

            fecha_actual = datetime.now().strftime("%Y-%m-%d")
            total_efectivo = 0
            total_tarjeta = 0

            for row in sheet.iter_rows(min_row=2, values_only=True):
                fecha_venta, _, _, total_venta, metodo_pago = row
                if fecha_venta == fecha_actual:
                    if metodo_pago == "efectivo":
                        total_efectivo += total_venta
                    elif metodo_pago == "tarjeta":
                        total_tarjeta += total_venta

            total_caja = dinero_inicial_caja + total_efectivo
            messagebox.showinfo("Corte de Caja", f"Total Efectivo: ${total_efectivo:.2f}\nTotal Tarjeta: ${total_tarjeta:.2f}\nTotal en Caja: ${total_caja:.2f}")

        except FileNotFoundError:
            messagebox.showerror("Error", "Archivo de ventas no encontrado.")

    boton_corte_caja = ctk.CTkButton(frame, text="Corte de Caja", command=corte_de_caja, image=icono_corte_caja, compound="left")
    boton_corte_caja.image = icono_corte_caja
    boton_corte_caja.grid(column=1, row=6, sticky=ctk.E)

    frame.grid_rowconfigure(4, weight=0)
    frame.grid_rowconfigure(5, weight=0)
    frame.grid_rowconfigure(6, weight=0)

    for child in frame.winfo_children():
        child.grid_configure(padx=5, pady=5)

    ventana_cobro.columnconfigure(0, weight=1)
    ventana_cobro.rowconfigure(1, weight=1)


# Inicio del programa
if __name__ == "__main__":
    cargar_ventas()
    cargar_productos()
    mostrar_ventana_login()