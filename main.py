import os
import openpyxl
import getpass
from datetime import datetime

# Variable para llevar un registro del dinero inicial en caja
dinero_inicial_caja = 0


# Función para el inicio de sesión
def login():
    while True:
        print("Por favor, inicia sesión:")
        username = input("Nombre de usuario: ")
        password = getpass.getpass("Contraseña: ")

        if username == "admin" and password == "adminpass":
            return "admin"
        elif username == "empleado" and password == "empleadopass":
            return "empleado"
        else:
            print("Credenciales incorrectas. Inténtalo de nuevo.")

# Función para agregar un producto
def agregar_producto(productos, tipo_usuario):
    if tipo_usuario == "admin":
        nombre = input("Nombre del producto: ")
        precio = float(input("Precio de venta: "))
        cantidad = int(input("Cantidad en inventario: "))

        producto = {"nombre": nombre, "precio": precio, "cantidad": cantidad}
        productos.append(producto)

        print(f"{nombre} ha sido agregado al inventario.")
        guardar_productos(productos)
    else:
        print("Acceso denegado. Debes ser admin para agregar productos.")

# Función para eliminar un producto
def eliminar_producto(productos, tipo_usuario):
    if tipo_usuario == "admin":
        if not productos:
            print("No hay productos para eliminar.")
            return

        print("Productos disponibles:")
        for i, producto in enumerate(productos):
            print(f"{i + 1}. {producto['nombre']}")

        choice = int(input("Selecciona el número del producto que deseas eliminar: ")) - 1

        if 0 <= choice < len(productos):
            producto_eliminado = productos.pop(choice)
            print(f"{producto_eliminado['nombre']} ha sido eliminado del inventario.")
            guardar_productos(productos)
        else:
            print("Selección inválida.")
    else:
        print("Acceso denegado. Debes ser admin para eliminar productos.")

# Función para modificar un producto
def modificar_producto(productos, tipo_usuario):
    if tipo_usuario == "admin":
        if not productos:
            print("No hay productos para modificar.")
            return

        print("Productos disponibles:")
        for i, producto in enumerate(productos):
            print(f"{i + 1}. {producto['nombre']}")

        choice = int(input("Selecciona el número del producto que deseas modificar: ")) - 1

        if 0 <= choice < len(productos):
            producto = productos[choice]
            print("Datos actuales del producto:")
            print(f"Nombre: {producto['nombre']}")
            print(f"Precio de venta: {producto['precio']}")
            print(f"Cantidad en inventario: {producto['cantidad']}")

            nombre = input("Nuevo nombre (dejar en blanco para no cambiar): ")
            if nombre:
                producto['nombre'] = nombre

            precio = input("Nuevo precio de venta (dejar en blanco para no cambiar): ")
            if precio:
                producto['precio'] = float(precio)

            cantidad = input("Nueva cantidad en inventario (dejar en blanco para no cambiar): ")
            if cantidad:
                producto['cantidad'] = int(cantidad)

            print(f"{producto['nombre']} ha sido modificado en el inventario.")
            guardar_productos(productos)
        else:
            print("Selección inválida.")
    else:
        print("Acceso denegado. Debes ser admin para modificar productos.")

# Función para la sección de caja de cobro
def caja_de_cobro(productos, tipo_usuario, ventas, total_ventas):
    caja = []

    while True:
        print("\nOpciones de caja de cobro:")
        print("1. Agregar producto a la caja")
        print("2. Eliminar producto de la caja")
        print("3. Hacer corte de caja")
        print("4. Cobrar productos y volver a la caja")
        print("5. Salir de la caja")

        opcion = input("Selecciona una opción: ")

        if opcion == "1":
            print("Productos disponibles:")
            for i, producto in enumerate(productos):
                print(f"{i + 1}. {producto['nombre']} - ${producto['precio']}")

            choice = int(input("Selecciona el número del producto que deseas agregar a la caja: ")) - 1

            if 0 <= choice < len(productos):
                producto = productos[choice]
                caja.append(producto)
                print(f"{producto['nombre']} ha sido agregado a la caja.")
            else:
                print("Selección inválida.")
        elif opcion == "2":
            if not caja:
                print("La caja está vacía.")
            else:
                print("Productos en la caja:")
                for i, producto in enumerate(caja):
                    print(f"{i + 1}. {producto['nombre']} - ${producto['precio']}")
                choice = int(input("Selecciona el número del producto que deseas eliminar de la caja: ")) - 1
                if 0 <= choice < len(caja):
                    producto_eliminado = caja.pop(choice)
                    print(f"{producto_eliminado['nombre']} ha sido eliminado de la caja.")
                else:
                    print("Selección inválida.")
        elif opcion == "3":
             total_ventas = mostrar_ventas_dia(ventas)  # Actualiza total_ventas con el valor calculado
             dinero_en_caja = float(input("Ingrese la cantidad de dinero en caja: $"))
             diferencia = dinero_en_caja - total_ventas
             print(f"Diferencia entre caja y ventas: ${diferencia:.2f}")

        elif opcion == "4":
            if not caja:
                print("La caja está vacía. No se puede realizar el cobro.")
            else:
                total_venta, metodo_pago = cobrar(caja)
                guardar_venta(total_venta, metodo_pago, caja, ventas)
                total_ventas += total_venta  # Actualizar el dinero total de las ventas
                caja = []  # Limpiar la caja después del cobro
        elif opcion == "5":
            break
        else:
            print("Opción inválida.")


# Función para mostrar las ventas del día
def mostrar_ventas_dia(ventas):
    fecha_actual = datetime.now().strftime("%Y-%m-%d")
    print(f"Ventas del día ({fecha_actual}):")
    ventas_del_dia = cargar_ventas(fecha_actual)

    if not ventas_del_dia:
        print("No hay ventas registradas para el día de hoy.")
        return

    total_ventas_dia = sum(venta[3] for venta in ventas_del_dia)  # Calcula el total de las ventas del día

    for venta in ventas_del_dia:
        print(f"ID Venta: {venta[1]}")
        print(f"Productos: {venta[2]}")
        print(f"Total Venta: ${venta[3]}")
        print(f"Método de Pago: {venta[4]}")
        print()

    print(f"Dinero total de las ventas: ${total_ventas_dia:.2f}")

    return total_ventas_dia  # Devuelve el total de las ventas del día

# Función para cargar las ventas de una fecha específica
def cargar_ventas(fecha):
    try:
        archivo_ventas = "ventas_totales.xlsx"
        if not os.path.exists(archivo_ventas):
            return []

        workbook = openpyxl.load_workbook(archivo_ventas)
        sheet = workbook.active
        ventas = []

        for row in sheet.iter_rows(min_row=2, values_only=True):
            if row[0] == fecha:
                ventas.append(row)

        return ventas
    except FileNotFoundError:
        return []

# Función para cobrar productos y calcular el cambio
def cobrar(caja):
    if not caja:
        print("La caja está vacía. No se puede realizar el cobro.")
        return 0, "n/a"

    total = sum(producto["precio"] for producto in caja)
    print(f"Total a cobrar: ${total}")
    metodo_pago = input("Método de pago (efectivo/tarjeta): ")

    if metodo_pago == "efectivo":
        pago_efectivo = float(input("Monto en efectivo: "))
        cambio = pago_efectivo - total

        if cambio < 0:
            print("El monto en efectivo es insuficiente.")
        else:
            print(f"¡Cambio: ${cambio:.2f}")
        return total, "efectivo"
    elif metodo_pago == "tarjeta":
        print("Por favor, verifique que el cobro se haya efectuado de manera correcta.")
        return total, "tarjeta"
    else:
        print("Método de pago inválido.")
        return 0, "n/a"

# Función para guardar los productos en un archivo Excel
def guardar_productos(productos):
    workbook = openpyxl.Workbook()
    sheet = workbook.active

    # Encabezados
    sheet.append(["Nombre", "Precio de Venta", "Cantidad en Inventario"])

    # Agrega los productos al archivo Excel
    for producto in productos:
        sheet.append([producto["nombre"], producto["precio"], producto["cantidad"]])

    workbook.save("productos.xlsx")

# Función para cargar productos desde un archivo Excel
def cargar_productos():
    try:
        workbook = openpyxl.load_workbook("productos.xlsx")
        sheet = workbook.active
        productos = []

        for row in sheet.iter_rows(min_row=2, values_only=True):
            nombre, precio, cantidad = row
            producto = {"nombre": nombre, "precio": precio, "cantidad": cantidad}
            productos.append(producto)

        return productos
    except FileNotFoundError:
        return []

# Función para guardar la venta en un archivo Excel
def guardar_venta(total, metodo_pago, caja, ventas):
    if not caja:
        print("No se puede guardar una venta vacía.")
        return

    fecha_actual = datetime.now().strftime("%Y-%m-%d")
    archivo_ventas = f"ventas_totales.xlsx"

    if not os.path.exists(archivo_ventas):
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.append(["Fecha", "ID Venta", "Productos", "Total Venta", "Método de Pago"])
    else:
        workbook = openpyxl.load_workbook(archivo_ventas)
        sheet = workbook.active

    productos_vendidos = ", ".join(producto["nombre"] for producto in caja)
    nueva_fila = [fecha_actual, len(sheet["A"]) + 1, productos_vendidos, total, metodo_pago]
    sheet.append(nueva_fila)

    workbook.save(archivo_ventas)

    ventas.append(nueva_fila)  # Agregar la venta al registro de ventas

if __name__ == "__main__":
    tipo_usuario = login()
    productos = cargar_productos()
    ventas = []  # Lista para mantener el registro de ventas durante la sesión
    total_ventas = 0  # Inicializa la variable total_ventas

    if tipo_usuario == "admin":
        while True:
            print("\nOpciones:")
            print("1. Agregar producto")
            print("2. Eliminar producto")
            print("3. Modificar producto")
            print("4. Caja de cobro")
            print("5. Guardar cambios y salir")

            opcion = input("Selecciona una opción: ")

            if opcion == "1":
                agregar_producto(productos, tipo_usuario)
            elif opcion == "2":
                eliminar_producto(productos, tipo_usuario)
            elif opcion == "3":
                modificar_producto(productos, tipo_usuario)
            elif opcion == "4":
                caja_de_cobro(productos, tipo_usuario, ventas, total_ventas)
            elif opcion == "5":
                guardar_productos(productos)
                break
            else:
                print("Opción inválida.")
    elif tipo_usuario == "empleado":
        while True:
            print("\nOpciones:")
            print("1. Caja de cobro")
            print("2. Salir")

            opcion = input("Selecciona una opción: ")

            if opcion == "1":
                caja_de_cobro(productos, tipo_usuario, ventas, total_ventas)
            elif opcion == "2":
                break
            else:
                print("Opción inválida.")
    else:
        print("Acceso denegado. Hasta luego.")
